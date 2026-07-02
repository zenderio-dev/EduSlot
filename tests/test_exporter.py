import csv
import json

from openpyxl import load_workbook

from eduslot import generate_schedule
from eduslot.exporter import (
    export_schedule_to_csv,
    export_schedule_to_json,
    export_schedule_to_xlsx,
)
from eduslot.models import (
    GroupLoad,
    LessonLoad,
    PreferencesInput,
    ScheduleItem,
    ScheduleResult,
    TeacherPreferenceInput,
    WorkloadInput,
)


def test_public_api_exposes_generate_schedule():
    workload = WorkloadInput(
        groups=[
            GroupLoad(
                name="Group A",
                lessons=[
                    LessonLoad(
                        subject="Python",
                        teacher="Ivanov",
                        lessons_per_week=1,
                    )
                ],
            )
        ]
    )
    preferences = PreferencesInput(
        preferences=[
            TeacherPreferenceInput(
                teacher="Ivanov",
                text="понедельник",
            )
        ]
    )

    result = generate_schedule(workload, preferences)

    assert result.conflicts == []
    assert len(result.schedule) == 1


def test_export_schedule_to_json_exports_full_schedule_result(tmp_path):
    result = ScheduleResult(
        schedule=[
            ScheduleItem(
                group="Group A",
                subject="Python",
                teacher="Ivanov",
                day="mon",
                slot=1,
                lesson_type="practice",
            )
        ],
        warnings=["demo warning"],
        conflicts=[],
    )
    output_path = tmp_path / "schedule.json"

    returned_path = export_schedule_to_json(result, output_path)

    assert returned_path == output_path
    assert output_path.exists()

    data = json.loads(output_path.read_text(encoding="utf-8"))

    assert data["schedule"][0]["group"] == "Group A"
    assert data["schedule"][0]["subject"] == "Python"
    assert data["warnings"] == ["demo warning"]
    assert data["conflicts"] == []


def test_export_schedule_to_json_accepts_schedule_items(tmp_path):
    schedule = [
        ScheduleItem(
            group="Group A",
            subject="Python",
            teacher="Ivanov",
            day="mon",
            slot=1,
            lesson_type="practice",
        )
    ]
    output_path = tmp_path / "schedule.json"

    export_schedule_to_json(schedule, output_path)

    data = json.loads(output_path.read_text(encoding="utf-8"))

    assert data == {
        "schedule": [
            {
                "group": "Group A",
                "subject": "Python",
                "teacher": "Ivanov",
                "day": "mon",
                "slot": 1,
                "lesson_type": "practice",
            }
        ]
    }


def test_export_schedule_to_csv(tmp_path):
    result = ScheduleResult(
        schedule=[
            ScheduleItem(
                group="Group A",
                subject="Python",
                teacher="Ivanov",
                day="mon",
                slot=1,
                lesson_type="practice",
            )
        ]
    )
    output_path = tmp_path / "schedule.csv"

    returned_path = export_schedule_to_csv(result, output_path)

    assert returned_path == output_path
    assert output_path.exists()

    with output_path.open(encoding="utf-8", newline="") as file:
        rows = list(csv.DictReader(file))

    assert rows == [
        {
            "group": "Group A",
            "subject": "Python",
            "teacher": "Ivanov",
            "day": "mon",
            "slot": "1",
            "lesson_type": "practice",
        }
    ]


def test_export_schedule_to_xlsx(tmp_path):
    result = ScheduleResult(
        schedule=[
            ScheduleItem(
                group="Group A",
                subject="Python",
                teacher="Ivanov",
                day="mon",
                slot=1,
                lesson_type="practice",
            )
        ]
    )
    output_path = tmp_path / "schedule.xlsx"

    returned_path = export_schedule_to_xlsx(result, output_path)

    assert returned_path == output_path
    assert output_path.exists()

    workbook = load_workbook(output_path)
    worksheet = workbook["Schedule"]

    rows = list(worksheet.iter_rows(values_only=True))

    assert rows == [
        ("group", "subject", "teacher", "day", "slot", "lesson_type"),
        ("Group A", "Python", "Ivanov", "mon", 1, "practice"),
    ]


def test_exporters_create_parent_directory(tmp_path):
    result = ScheduleResult(
        schedule=[
            ScheduleItem(
                group="Group A",
                subject="Python",
                teacher="Ivanov",
                day="mon",
                slot=1,
                lesson_type="practice",
            )
        ]
    )
    output_path = tmp_path / "nested" / "exports" / "schedule.json"

    export_schedule_to_json(result, output_path)

    assert output_path.exists()